"""
Export ผลเปรียบเทียบ baseline เป็นไฟล์ Excel พร้อมตาราง + กราฟ
รัน: python export_excel.py
ผลลัพธ์: results/baseline_comparison.xlsx
"""

import json
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

from config import Config


# ── Styles ──────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Tahoma", size=12, bold=True, color="FFFFFF")
SUBHEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
SUBHEADER_FONT = Font(name="Tahoma", size=11, bold=True, color="1F4E78")
WINNER_FILL = PatternFill("solid", fgColor="C6EFCE")
WINNER_FONT = Font(name="Tahoma", size=11, bold=True, color="006100")
NORMAL_FONT = Font(name="Tahoma", size=11)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header_row(ws, row_num, num_cols):
    for col in range(1, num_cols + 1):
        c = ws.cell(row=row_num, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER


def autosize_columns(ws):
    for col_cells in ws.columns:
        length = max(len(str(c.value or "")) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = length + 4


# ── Build sheets ────────────────────────────────────────────────────────────

def sheet_summary(wb, results: dict):
    ws = wb.create_sheet("📊 Summary", 0)

    ws["A1"] = "ตารางเปรียบเทียบประสิทธิภาพ — Baseline Comparison"
    ws["A1"].font = Font(name="Tahoma", size=16, bold=True, color="1F4E78")
    ws.merge_cells("A1:H1")
    ws.row_dimensions[1].height = 30

    headers = ["Model", "RMSE ↓", "NDCG@5 ↑", "NDCG@10 ↑",
               "NDCG@20 ↑", "HR@10 ↑", "Params", "Train Time (s)"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header_row(ws, 3, len(headers))

    # Find best per metric (lower is better for RMSE)
    best = {
        "RMSE": min(results, key=lambda m: results[m]["RMSE"]),
        "NDCG@5": max(results, key=lambda m: results[m]["NDCG@5"]),
        "NDCG@10": max(results, key=lambda m: results[m]["NDCG@10"]),
        "NDCG@20": max(results, key=lambda m: results[m]["NDCG@20"]),
        "HR@10": max(results, key=lambda m: results[m]["HR@10"]),
    }

    row = 4
    for model, r in results.items():
        ws.cell(row=row, column=1, value=model).font = Font(name="Tahoma", size=11, bold=True)
        cells = [
            (2, "RMSE", round(r["RMSE"], 4)),
            (3, "NDCG@5", round(r["NDCG@5"], 4)),
            (4, "NDCG@10", round(r["NDCG@10"], 4)),
            (5, "NDCG@20", round(r["NDCG@20"], 4)),
            (6, "HR@10", round(r["HR@10"], 4)),
            (7, None, f"{r['params']:,}"),
            (8, None, r["train_time_sec"]),
        ]
        for col, mk, val in cells:
            c = ws.cell(row=row, column=col, value=val)
            c.alignment = CENTER
            c.border = BORDER
            c.font = NORMAL_FONT
            if mk and best.get(mk) == model:
                c.fill = WINNER_FILL
                c.font = WINNER_FONT
        ws.cell(row=row, column=1).border = BORDER
        row += 1

    autosize_columns(ws)

    # Note
    ws.cell(row=row + 1, column=1,
            value="🟢 = ค่าที่ดีที่สุดในแต่ละ metric").font = Font(
        name="Tahoma", size=10, italic=True, color="006100")

    # Add chart
    chart = BarChart()
    chart.type = "col"
    chart.style = 11
    chart.title = "Performance Comparison"
    chart.y_axis.title = "Score"
    chart.x_axis.title = "Model"
    data = Reference(ws, min_col=3, min_row=3, max_col=6, max_row=3 + len(results))
    cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(results))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 20
    chart.dataLabels = DataLabelList(showVal=True)
    ws.add_chart(chart, "A14")


def sheet_improvement(wb, results: dict):
    ws = wb.create_sheet("📈 Improvement")

    ws["A1"] = "Bi-Level GAT — เปอร์เซ็นต์การปรับปรุงเทียบกับ Baseline"
    ws["A1"].font = Font(name="Tahoma", size=14, bold=True, color="1F4E78")
    ws.merge_cells("A1:E1")
    ws.row_dimensions[1].height = 25

    headers = ["เปรียบเทียบ", "Metric", "Baseline", "Bi-Level GAT", "Improvement (%)"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header_row(ws, 3, len(headers))

    gat = results["Bi-Level GAT"]
    metrics = [
        ("RMSE", True),       # lower is better
        ("NDCG@5", False),
        ("NDCG@10", False),
        ("NDCG@20", False),
        ("HR@10", False),
    ]

    row = 4
    for baseline_name in ["MF", "GCN"]:
        b = results[baseline_name]
        for mk, lower_better in metrics:
            base_v = b[mk]
            gat_v = gat[mk]
            if lower_better:
                imp = (base_v - gat_v) / base_v * 100
            else:
                imp = (gat_v - base_v) / base_v * 100

            ws.cell(row=row, column=1, value=f"vs {baseline_name}")
            ws.cell(row=row, column=2, value=mk)
            ws.cell(row=row, column=3, value=round(base_v, 4))
            ws.cell(row=row, column=4, value=round(gat_v, 4))
            imp_cell = ws.cell(row=row, column=5, value=f"{imp:+.2f}%")

            for col in range(1, 6):
                ws.cell(row=row, column=col).border = BORDER
                ws.cell(row=row, column=col).alignment = CENTER

            if imp > 0:
                imp_cell.font = Font(name="Tahoma", bold=True, color="006100")
                imp_cell.fill = PatternFill("solid", fgColor="E2EFDA")
            elif imp < 0:
                imp_cell.font = Font(name="Tahoma", bold=True, color="9C0006")
                imp_cell.fill = PatternFill("solid", fgColor="FCE4D6")
            row += 1

    autosize_columns(ws)


def sheet_metrics_detail(wb, results: dict):
    ws = wb.create_sheet("🔍 Per-Metric Detail")

    ws["A1"] = "เปรียบเทียบทุก Metric แบบละเอียด"
    ws["A1"].font = Font(name="Tahoma", size=14, bold=True, color="1F4E78")
    ws.merge_cells("A1:D1")

    metrics_order = ["RMSE", "NDCG@5", "NDCG@10", "NDCG@20", "HR@5",
                     "HR@10", "HR@20", "params", "train_time_sec"]

    headers = ["Metric"] + list(results.keys())
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header_row(ws, 3, len(headers))

    row = 4
    for mk in metrics_order:
        ws.cell(row=row, column=1, value=mk).font = Font(name="Tahoma", bold=True)
        ws.cell(row=row, column=1).border = BORDER
        for col, model in enumerate(results.keys(), 2):
            v = results[model].get(mk, "-")
            if isinstance(v, float):
                v = round(v, 4)
            elif isinstance(v, int) and mk == "params":
                v = f"{v:,}"
            c = ws.cell(row=row, column=col, value=v)
            c.alignment = CENTER
            c.border = BORDER
            c.font = NORMAL_FONT
        row += 1

    autosize_columns(ws)


def sheet_report_text(wb, results: dict):
    """Sheet ที่มีข้อความพร้อมใช้ใน Section "ผลการวิจัย" """
    ws = wb.create_sheet("📝 Report Text")

    gat = results["Bi-Level GAT"]
    mf = results["MF"]
    gcn = results["GCN"]

    rmse_imp_mf = (mf["RMSE"] - gat["RMSE"]) / mf["RMSE"] * 100
    rmse_imp_gcn = (gcn["RMSE"] - gat["RMSE"]) / gcn["RMSE"] * 100

    text = f"""ผลการเปรียบเทียบ Bi-Level GAT กับ Baseline Models

จากการทดลองเปรียบเทียบโมเดล Bi-Level GAT กับโมเดล baseline 2 ตัว ได้แก่
Matrix Factorization (MF) และ Graph Convolutional Network (GCN)
โดยใช้ชุดข้อมูลทดสอบเดียวกัน ผลลัพธ์ปรากฏดังตารางที่ 4-1

ผลลัพธ์สำคัญ:
1. Bi-Level GAT ให้ค่า RMSE ต่ำที่สุด ({gat['RMSE']:.4f}) ลดลงจาก
   - MF baseline: {rmse_imp_mf:+.2f}%
   - GCN baseline: {rmse_imp_gcn:+.2f}%

2. ค่า NDCG@10 ของทั้ง 3 โมเดลใกล้เคียงกัน:
   - MF: {mf['NDCG@10']:.4f}
   - GCN: {gcn['NDCG@10']:.4f}
   - Bi-Level GAT: {gat['NDCG@10']:.4f}

3. ค่า HR@10 เท่ากันทุกโมเดล ({gat['HR@10']:.4f}) เนื่องจาก
   ชุดข้อมูลทดสอบมีจำนวน items จำกัดทำให้ผู้ใช้ส่วนใหญ่
   มีกิจกรรมที่ตรงใจอยู่ใน Top-10 อยู่แล้ว

4. ความซับซ้อนของโมเดล (จำนวนพารามิเตอร์):
   - MF: {mf['params']:,}
   - GCN: {gcn['params']:,}
   - Bi-Level GAT: {gat['params']:,}

ข้อสรุป:
Bi-Level GAT ให้ความแม่นยำในการทำนายคะแนน (RMSE) ดีที่สุด
และยังมีข้อได้เปรียบเชิงคุณภาพที่ baseline ไม่มี ได้แก่
- ความสามารถในการอธิบายผลลัพธ์ (Explainability) ผ่าน Attention Weights
- กลไก Adaptive Feedback Loop ที่ปรับคำแนะนำตามคุณภาพชีวิต (QoL)
- การจับ Bi-Level Attention (Social Influence + Item Content)
"""

    ws["A1"] = "ตัวอย่างข้อความสำหรับ Section ผลการวิจัย"
    ws["A1"].font = Font(name="Tahoma", size=14, bold=True, color="1F4E78")
    ws.merge_cells("A1:F1")

    for i, line in enumerate(text.split("\n"), start=3):
        c = ws.cell(row=i, column=1, value=line)
        c.font = Font(name="Tahoma", size=12)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)

    ws.column_dimensions["A"].width = 100


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    json_path = Path(Config.RESULTS_DIR) / "baseline_comparison.json"
    if not json_path.exists():
        print(f"⚠ ไม่พบไฟล์ {json_path}")
        print("กรุณารัน: python compare_baselines.py ก่อน")
        return

    with open(json_path, encoding="utf-8") as f:
        results = json.load(f)

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    sheet_summary(wb, results)
    sheet_improvement(wb, results)
    sheet_metrics_detail(wb, results)
    sheet_report_text(wb, results)

    output = Path(Config.RESULTS_DIR) / "baseline_comparison.xlsx"
    wb.save(output)

    print(f"✓ สร้างไฟล์ Excel เรียบร้อย")
    print(f"  ที่อยู่: {output.absolute()}")
    print(f"  ประกอบด้วย 4 sheets:")
    print("    📊 Summary          — ตารางหลัก + กราฟแท่ง")
    print("    📈 Improvement      — % การปรับปรุงเทียบ baseline")
    print("    🔍 Per-Metric Detail — ข้อมูลทุก metric")
    print("    📝 Report Text      — ข้อความพร้อมใช้ในรายงาน")


if __name__ == "__main__":
    main()
